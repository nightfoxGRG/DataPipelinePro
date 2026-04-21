"""Generate an Excel configuration file in tables_config_v2 format
from a list of inferred column descriptors.
"""

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


_TEMPLATE_PATH = Path(__file__).parent.parent / 'static' / 'TablesConfig.xlsm'

# Number of data columns in one table block (matches the template header row)
_V2_DATA_COLS = 9


def generate_excel_config_v2(table_name: str, columns: list[dict]) -> bytes:
    """Build and return bytes of an xlsm workbook containing a tables_config_v2 sheet.

    Loads TablesConfig.xlsm as a template and fills in the first table block
    without deleting any rows or overwriting styles, formulas, or other settings.
    Only cell values are written.

    Each dict in *columns* must contain at least:
      'code'    – column code (SQL identifier)
      'db_type' – PostgreSQL type
    Optional keys: 'label', 'size'.
    """
    wb = load_workbook(_TEMPLATE_PATH, keep_vba=True)

    # ----- Clear sample data values from tables_config (preserve row structure) -----
    if 'tables_config' in wb.sheetnames:
        tc = wb['tables_config']
        for row in tc.iter_rows(min_row=1, max_row=tc.max_row, min_col=2, max_col=tc.max_column):
            for cell in row:
                cell.value = None

    ws = wb['tables_config_v2']

    # ----- Write table name into B1 (A1 already holds "Наименование таблицы") -----
    ws.cell(row=1, column=2).value = table_name

    # ----- Clear old sample data values in block-1 data rows (rows 3+, cols A-I) -----
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=_V2_DATA_COLS):
        for cell in row:
            cell.value = None

    # ----- Write column data starting at row 3 -----
    for row_idx, col_info in enumerate(columns, start=3):
        ws.cell(row=row_idx, column=1).value = col_info.get('label') or col_info['code']
        ws.cell(row=row_idx, column=2).value = col_info['code']
        ws.cell(row=row_idx, column=3).value = col_info['db_type']
        size = col_info.get('size')
        if size:
            ws.cell(row=row_idx, column=4).value = size

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
