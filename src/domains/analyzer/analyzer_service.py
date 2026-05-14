# analyzer_service.py
"""Оркестрация анализа таблиц рабочей БД на подобные дубликаты.

Читает данные чанками, передаёт в AnalyzerDuplicateService,
формирует Excel-отчёт: отдельная закладка на каждую таблицу.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from common.context_service import ContextService
from common.error import AppError
from common.singleton_meta import SingletonMeta
from domains.analyzer.analyzer_duplicate_service import AnalyzerDuplicateService
from domains.working_db.information_schema_repository import InformationSchemaRepository
from domains.working_db.working_db_repository import WorkingDbRepository

_HEADER_FILL = PatternFill('solid', fgColor='D9E1F2')
_HEADER_FONT = Font(bold=True, size=11)
_GROUP_FILL = PatternFill('solid', fgColor='FCE4D6')
_GROUP_FONT = Font(bold=True, size=10, color='C00000')


class AnalyzerService(metaclass=SingletonMeta):

    def __init__(self) -> None:
        self._working_db_repository = WorkingDbRepository()
        self._information_schema_repository = InformationSchemaRepository()
        self._duplicate_service = AnalyzerDuplicateService()

    def get_tables(self) -> list[str]:
        user = ContextService.get_user_info()
        if not user.db_id or not user.project_schema:
            raise AppError('Рабочая БД или схема не определены.')
        return self._information_schema_repository.get_base_tables(user.db_id, user.project_schema)

    def analyze_tables(self, tables: list[str], chunk_size: int) -> bytes:
        user = ContextService.get_user_info()
        if not user.db_id or not user.project_schema:
            raise AppError('Рабочая БД или схема не определены.')
        tables = [t for t in tables if t]
        if not tables:
            raise AppError('Не выбраны таблицы для анализа.')
        if chunk_size < 1:
            chunk_size = 1000

        self._duplicate_service.reset()

        for table in tables:
            self._process_table(user.db_id, user.project_schema, table, chunk_size)

        return self._build_excel(tables)

    def _process_table(self, db_id: int, schema: str, table: str, chunk_size: int) -> None:
        columns = self._working_db_repository.get_column_names(db_id, schema, table)
        if not columns:
            return
        offset = 0
        while True:
            chunk = self._working_db_repository.fetch_chunk(db_id, schema, table, offset, chunk_size)
            if not chunk:
                break
            self._duplicate_service.process_chunk(table, columns, chunk)
            offset += chunk_size
            if len(chunk) < chunk_size:
                break

    def _build_excel(self, tables: list[str]) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)  # убираем дефолтный лист

        for table in tables:
            duplicates = self._duplicate_service.get_duplicates(table)
            sheet_name = table[:31]  # Excel ограничивает имя листа 31 символом
            ws = wb.create_sheet(title=sheet_name)

            if not duplicates:
                ws.append(['Подобных дубликатов не обнаружено'])
                ws.cell(1, 1).font = Font(italic=True, color='808080')
                ws.column_dimensions['A'].width = 40
                continue

            row = 1
            for col_name, groups in sorted(duplicates.items()):
                # заголовок колонки
                header_cell = ws.cell(row=row, column=1, value=f'Подобный дубликат — {col_name}')
                header_cell.font = _HEADER_FONT
                header_cell.fill = _HEADER_FILL
                header_cell.alignment = Alignment(wrap_text=False)
                row += 1

                for group in groups:
                    for value in group:
                        cell = ws.cell(row=row, column=1, value=value)
                        cell.fill = _GROUP_FILL
                        cell.font = _GROUP_FONT
                        row += 1
                    # пустая строка между группами
                    row += 1

                # пустая строка между полями
                row += 1

            ws.column_dimensions['A'].width = 60

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
