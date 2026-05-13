# loader_by_directory_service.py
"""Загрузка данных из директории с явным выбором конфигурации маппинга."""
import csv
from io import BytesIO, StringIO
from pathlib import Path

from flask import Request, Response, jsonify
from openpyxl import load_workbook

from common.context_service import ContextService
from common.error import AppError
from common.singleton_meta import SingletonMeta
from domains.loader.data_stream import DataStream
from domains.loader.loader_service import LoaderService
from domains.source_to_table.source_to_table_config_repository import SourceToTableConfigRepository

_ALLOWED_EXTENSIONS = {'.xlsx', '.xlsm', '.csv'}


class LoaderByDirectoryService(metaclass=SingletonMeta):

    def __init__(self) -> None:
        self._config_repository = SourceToTableConfigRepository()
        self._loader_service = LoaderService()

    def list_configs(self) -> Response:
        user = ContextService.get_user_info()
        if not user.project_id:
            raise AppError('Проект не определён.')
        configs = self._config_repository.find_by_project(user.project_id)
        return jsonify(configs=[
            {
                'id': c.id,
                'table_name': c.table_name,
                'code': c.code,
                'description': c.description or '',
            }
            for c in configs
        ])

    def load_from_directory(self, request: Request) -> Response:
        config_id_raw = request.form.get('config_id', '')
        if not config_id_raw:
            raise AppError('Не выбрана конфигурация маппинга.')
        try:
            config_id = int(config_id_raw)
        except (ValueError, TypeError):
            raise AppError('Некорректный идентификатор конфигурации.')

        files = [f for f in request.files.getlist('data_files') if f and f.filename]
        if not files:
            raise AppError('Не выбраны файлы для загрузки.')

        user = ContextService.get_user_info()
        if not user.project_id:
            raise AppError('Проект не определён.')
        if not user.db_id:
            raise AppError('Рабочая БД не определена.')
        if not user.project_schema:
            raise AppError('Схема проекта не определена.')

        config = self._config_repository.find_by_id_and_project(config_id, user.project_id)
        if not config:
            raise AppError('Конфигурация маппинга не найдена.')

        results: list[dict] = []
        for file_storage in files:
            source_name = Path(file_storage.filename).name
            result = self._process_file(
                file_storage=file_storage,
                source_name=source_name,
                config_id=config_id,
                db_id=user.db_id,
                schema=user.project_schema,
            )
            results.append(result)

        return jsonify(results=results)

    def _process_file(self, file_storage, source_name: str, config_id: int, db_id: int, schema: str) -> dict:
        extension = Path(source_name).suffix.lower()
        if extension not in _ALLOWED_EXTENSIONS:
            return {'file': source_name, 'success': False, 'reason': 'неподдерживаемый формат файла'}

        content = file_storage.read()
        try:
            count = self._load_content(content, extension, config_id, db_id, schema, source_name)
        except AppError as exc:
            return {'file': source_name, 'success': False, 'reason': exc.errors[0] if exc.errors else str(exc)}
        except Exception as exc:  # noqa: BLE001
            return {'file': source_name, 'success': False, 'reason': str(exc)}

        return {'file': source_name, 'success': True, 'count': count}

    def _load_content(self, content: bytes, extension: str, config_id: int, db_id: int, schema: str, source_name: str) -> int:
        if extension in {'.xlsx', '.xlsm'}:
            headers = self._read_excel_headers(content)
            stream = DataStream(headers=headers, rows_factory=lambda: self._excel_rows(content))
            return self._loader_service.load(db_id, schema, config_id, source_name, stream)

        text = self._decode_csv(content)
        headers = self._read_csv_headers(text)
        stream = DataStream(headers=headers, rows_factory=lambda: self._csv_rows(text))
        return self._loader_service.load(db_id, schema, config_id, source_name, stream)

    @staticmethod
    def _read_excel_headers(content: bytes) -> list[str]:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row is None:
                raise AppError('Файл пустой.')
            return [str(c) if c is not None else '' for c in first_row]
        finally:
            wb.close()

    @staticmethod
    def _excel_rows(content: bytes):
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter, None)
            for row in rows_iter:
                yield row
        finally:
            wb.close()

    @staticmethod
    def _read_csv_headers(text: str) -> list[str]:
        reader = csv.reader(StringIO(text))
        first_row = next(reader, None)
        if first_row is None:
            raise AppError('Файл пустой.')
        return list(first_row)

    @staticmethod
    def _csv_rows(text: str):
        reader = csv.reader(StringIO(text))
        next(reader, None)
        for row in reader:
            yield row

    @staticmethod
    def _decode_csv(content: bytes) -> str:
        for encoding in ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1'):
            try:
                return content.decode(encoding)
            except (UnicodeDecodeError, LookupError):
                continue
        return content.decode('latin-1')
