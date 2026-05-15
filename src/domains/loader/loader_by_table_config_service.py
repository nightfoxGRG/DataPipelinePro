# loader_by_table_config_service.py
"""Загрузка данных из директории файлов по конфигурационному файлу проекта.

Для каждого файла из директории ищется таблица в табличной конфигурации
проекта по 'original_name' (имя файла в строке "Наименование таблицы").
Если таблица найдена и есть source_to_table_config с is_auto_generated=True
— запускается LoaderService.load.
"""
import csv
from io import BytesIO, StringIO
from pathlib import Path

from flask import Request, Response, jsonify
from openpyxl import load_workbook

from common.context_service import ContextService
from common.error import AppError
from common.singleton_meta import SingletonMeta
from domains.configurator.table_config_generator_service import TABLE_CONFIG_BUCKET
from domains.configurator.table_config_parser_service import TableConfigParserService
from domains.loader.data_stream import DataStream
from domains.loader.loader_service import LoaderService
from domains.minio.minio_service import MinioService
from domains.project.project_repository import ProjectRepository
from domains.source_to_table.source_to_table_config_repository import SourceToTableConfigRepository

_ALLOWED_EXTENSIONS = {'.xlsx', '.xlsm', '.csv'}


class LoaderByTableConfigService(metaclass=SingletonMeta):

    def __init__(self) -> None:
        self._parser = TableConfigParserService()
        self._minio = MinioService()
        self._project_repository = ProjectRepository()
        self._config_repository = SourceToTableConfigRepository()
        self._loader_service = LoaderService()

    def load_from_directory(self, request: Request) -> Response:
        files = [f for f in request.files.getlist('data_files') if f and f.filename]
        if not files:
            raise AppError('Не выбраны файлы для загрузки.')

        user = ContextService.get_user_info()
        if not user.project_id:
            raise AppError('Проект не определён.')
        if not user.db_id:
            raise AppError('Рабочая БД не определена.')
        if not user.project_schema:
            raise AppError('Схема проекта не определена.')

        project = self._project_repository.find_by_id(user.project_id)
        if not project or not project.table_config_minio_id:
            raise AppError('Конфигурационный файл в системе отсутствует.')

        config_content = self._minio.download_bytes(TABLE_CONFIG_BUCKET, project.table_config_minio_id)
        tables = self._parser.parse_tables_config(config_content, 'config.xlsm')

        # filename (lowercase, basename with extension) -> table_name
        filename_to_table: dict[str, str] = {}
        for table in tables:
            if table.original_name:
                filename_to_table[Path(table.original_name).name.lower()] = table.name

        results: list[dict] = []
        for file_storage in files:
            source_name = Path(file_storage.filename).name
            result = self._process_file(
                file_storage=file_storage,
                source_name=source_name,
                filename_to_table=filename_to_table,
                project_id=user.project_id,
                db_id=user.db_id,
                schema=user.project_schema,
                user_id=user.user_id,
            )
            results.append(result)

        return jsonify(results=results)

    def _process_file(
        self,
        file_storage,
        source_name: str,
        filename_to_table: dict[str, str],
        project_id: int,
        db_id: int,
        schema: str,
        user_id: int,
    ) -> dict:
        extension = Path(source_name).suffix.lower()
        if extension not in _ALLOWED_EXTENSIONS:
            return {'file': source_name, 'success': False, 'reason': 'неподдерживаемый формат файла'}

        table_name = filename_to_table.get(source_name.lower())
        if not table_name:
            return {'file': source_name, 'success': False, 'reason': 'таблица для файла не найдена в конфигурации'}

        configs = self._config_repository.find_by_project_and_table(project_id, table_name)
        config = next((c for c in configs if c.is_auto_generated), None)
        if not config:
            return {
                'file': source_name,
                'success': False,
                'reason': f'автоматически сгенерированный маппинг не найден для таблицы {table_name}',
            }

        content = file_storage.read()
        try:
            count = self._load_content(
                content=content,
                extension=extension,
                config_id=config.id,
                db_id=db_id,
                schema=schema,
                source_name=source_name,
            )
        except AppError as exc:
            return {
                'file': source_name,
                'success': False,
                'reason': exc.errors[0] if exc.errors else str(exc),
            }
        except Exception as exc:  # noqa: BLE001 — каждый файл изолирован
            return {'file': source_name, 'success': False, 'reason': str(exc)}

        return {'file': source_name, 'success': True, 'count': count, 'table': table_name}

    def _load_content(
        self,
        content: bytes,
        extension: str,
        config_id: int,
        db_id: int,
        schema: str,
        source_name: str,
    ) -> int:
        if extension in {'.xlsx', '.xlsm'}:
            headers = self._read_excel_headers(content)
            stream = DataStream(headers=headers, rows_factory=lambda: self._excel_rows(content))
            return self._loader_service.load(db_id, schema, config_id, source_name, stream)

        # CSV
        text = self._decode_csv(content)
        dialect = self._sniff_dialect(text)
        headers = self._read_csv_headers(text, dialect)
        stream = DataStream(headers=headers, rows_factory=lambda: self._csv_rows(text, dialect))
        return self._loader_service.load(db_id, schema, config_id, source_name, stream)

    @staticmethod
    def _read_excel_headers(content: bytes) -> list[str]:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row is None:
                raise AppError('Файл пустой.')
            return [str(c) if c is not None else '' for c in first_row]
        finally:
            wb.close()

    @staticmethod
    def _excel_rows(content: bytes):
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter, None)  # пропускаем заголовок
            for row in rows_iter:
                yield row
        finally:
            wb.close()

    @staticmethod
    def _sniff_dialect(text: str):
        try:
            return csv.Sniffer().sniff(text[:4096], delimiters=',;\t|')
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ';' if ';' in text[:4096] else ','
            return dialect

    @staticmethod
    def _read_csv_headers(text: str, dialect) -> list[str]:
        reader = csv.reader(StringIO(text), dialect)
        first_row = next(reader, None)
        if first_row is None:
            raise AppError('Файл пустой.')
        return list(first_row)

    @staticmethod
    def _csv_rows(text: str, dialect):
        reader = csv.reader(StringIO(text), dialect)
        next(reader, None)  # пропускаем заголовок
        for row in reader:
            yield row

    @staticmethod
    def _decode_csv(content: bytes) -> str:
        for encoding in ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1'):
            try:
                return content.decode(encoding)
            except (UnicodeDecodeError, LookupError):
                continue
        return content.decode('latin-1')
