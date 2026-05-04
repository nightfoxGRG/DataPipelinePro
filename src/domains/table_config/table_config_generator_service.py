#table_config_generator_service.py
import re
from datetime import datetime
import zipfile
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from common.singleton_meta import SingletonMeta
from flask import Request, Response, g, jsonify
from openpyxl import load_workbook

from common.error import AppError, ValidationError
from common.project_paths import ProjectPaths
from config.db_orm_sqlalchemy.db_session_config import session_scope
from domains.minio.minio_service import MinioService
from domains.project.project_repository import ProjectRepository
from domains.table_config.table_config_data_file_reader_service import (
    ALLOWED_DATA_EXTENSIONS,
    TableConfigDataFileReaderService,
)
from domains.table_config.table_config_data_file_validator import TableConfigDataFileValidator
from domains.table_config.table_config_model import ColumnConfig, TableConfig
from domains.table_config.table_config_parser_service import TableConfigParserService
from domains.table_config.table_config_validator import TableConfigValidator
from utils.file_util import read_uploaded_file

_TEMPLATE_PATH = ProjectPaths.STATIC / 'TablesConfig.xlsm'
_V2_DATA_COLS = 9
_TC_DATA_ROWS = 10
_TC_BLOCK_STRIDE = 12
_TC_MAX_BLOCKS = 5
_TC_BUCKET = 'data-pipeline-table-config'
_TC_BUCKET_ARCH = 'data-pipeline-table-config-arch'

_AUTO_PK_COL = {'label': 'Ид', 'code': 'id', 'db_type': 'bigserial', 'primary_key': True}
_PACKAGE_ID_COL = {'label': 'Пакетный ид', 'code': 'package_id', 'db_type': 'varchar'}
_PACKAGE_TS_COL = {'label': 'Пакетный временной штамп', 'code': 'package_timestamp', 'db_type': 'timestamptz'}


class TableConfigGeneratorService(metaclass=SingletonMeta):

    def __init__(
        self,
        reader: TableConfigDataFileReaderService | None = None,
        validator: TableConfigDataFileValidator | None = None,
        minio: MinioService | None = None,
        config_validator: TableConfigValidator | None = None,
        config_parser: TableConfigParserService | None = None,
    ) -> None:
        self._reader = reader or TableConfigDataFileReaderService()
        self._validator = validator or TableConfigDataFileValidator()
        self._minio = minio or MinioService()
        self._config_validator = config_validator or TableConfigValidator()
        self._config_parser = config_parser or TableConfigParserService()

    @staticmethod
    def _to_table_configs(tables: list[tuple[str, str | None, list[dict]]]) -> list[TableConfig]:
        result = []
        for name, _, cols in tables:
            columns = [
                ColumnConfig(
                    name=c['code'],
                    db_type=c['db_type'],
                    size=c.get('size'),
                    nullable=c.get('nullable', True),
                    unique=c.get('unique', False),
                    primary_key=c.get('primary_key', False),
                    foreign_key=c.get('foreign_key'),
                    default=c.get('default'),
                    label=c.get('label'),
                )
                for c in cols
            ]
            result.append(TableConfig(name=name, columns=columns))
        return result

    def generate_table_config_from_data_file(self, request: Request) -> Response:
        content, filename = read_uploaded_file(request.files.get('data_file'), ALLOWED_DATA_EXTENSIONS)
        add_pk = request.form.get('add_pk') == '1'
        add_package_fields = request.form.get('add_package_fields') == '1'
        mode = request.form.get('mode') or None  # None | 'replace' | 'append'

        try:
            table_name, original_name, columns = self._read_and_infer(content, filename)
        except AppError:
            raise
        except Exception as exc:
            raise AppError(str(exc)) from exc

        return self._persist_and_respond([(table_name, original_name, columns)], add_pk, add_package_fields, mode)

    def generate_table_config_from_directory(self, request: Request) -> Response:
        files = request.files.getlist('data_files')
        files = [f for f in files if f and f.filename]
        if not files:
            raise AppError('Не выбраны файлы для загрузки.')
        if len(files) > _TC_MAX_BLOCKS:
            raise AppError(
                f'Превышено максимальное количество файлов: {len(files)} (допустимо {_TC_MAX_BLOCKS}).'
            )

        add_pk = request.form.get('add_pk') == '1'
        add_package_fields = request.form.get('add_package_fields') == '1'
        mode = request.form.get('mode') or None

        tables: list[tuple[str, str, list[dict]]] = []
        errors: list[str] = []
        for file_storage in files:
            display_name = Path(file_storage.filename).name
            try:
                content, filename = read_uploaded_file(file_storage, ALLOWED_DATA_EXTENSIONS)
                table_name, original_name, columns = self._read_and_infer(content, filename)
                tables.append((table_name, original_name, columns))
            except ValidationError as exc:
                errors.extend(f'{display_name}: {msg}' for msg in exc.errors)
            except AppError as exc:
                errors.append(f'{display_name}: {exc.errors[0] if exc.errors else str(exc)}')
            except Exception as exc:
                errors.append(f'{display_name}: {exc}')

        if errors:
            raise ValidationError(errors=errors)

        return self._persist_and_respond(tables, add_pk, add_package_fields, mode)

    def _persist_and_respond(
        self,
        tables: list[tuple[str, str, list[dict]]],
        add_pk: bool,
        add_package_fields: bool,
        mode: str | None,
    ) -> Response:
        current_user = getattr(g, 'current_user', None)
        project_id = getattr(current_user, 'project_id', None) if current_user else None
        project_schema = getattr(current_user, 'project_schema', None) if current_user else None

        self._config_validator.validate_tables(self._to_table_configs(tables))

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        object_name = f'{project_schema}_table_config_{ts}' if project_schema else f'table_config_{ts}'

        if not project_id:
            xlsx_bytes = self.generate_excel_config_multi(tables, add_pk, add_package_fields)
            self._minio.upload_bytes(
                _TC_BUCKET, object_name, xlsx_bytes,
                content_type='application/vnd.ms-excel.sheet.macroEnabled.12',
            )
            return jsonify(success=True)

        with session_scope() as session:
            projectRepository = ProjectRepository(session)
            project = projectRepository.find_by_id(project_id)
            existing_name = project.table_config_minio_id if project else None

            if existing_name and not mode:
                return jsonify(choose_mode=True)

            if mode == 'append' and existing_name:
                existing_bytes = self._minio.download_bytes(_TC_BUCKET, existing_name)
                xlsx_bytes = self._append_tables(existing_bytes, tables, add_pk, add_package_fields)
            else:
                xlsx_bytes = self.generate_excel_config_multi(tables, add_pk, add_package_fields)

            if existing_name:
                try:
                    self._minio.copy_to_bucket(_TC_BUCKET, _TC_BUCKET_ARCH, existing_name)
                    self._minio.delete(_TC_BUCKET, existing_name)
                except AppError:
                    pass

            self._minio.upload_bytes(
                _TC_BUCKET, object_name, xlsx_bytes,
                content_type='application/vnd.ms-excel.sheet.macroEnabled.12',
            )

            if project:
                project.table_config_minio_id = object_name
                projectRepository.save(project)

        return jsonify(success=True)

    def upload_table_config_file(self, request: Request) -> Response:
        file = request.files.get('config_file')
        if not file or not file.filename:
            raise AppError('Файл не выбран.')
        content = file.read()
        if not content:
            raise AppError('Загруженный файл пуст.')

        current_user = getattr(g, 'current_user', None)
        project_id = getattr(current_user, 'project_id', None) if current_user else None
        project_schema = getattr(current_user, 'project_schema', None) if current_user else None

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        object_name = f'{project_schema}_table_config_{ts}' if project_schema else f'table_config_{ts}'

        if not project_id:
            self._minio.upload_bytes(
                _TC_BUCKET, object_name, content,
                content_type='application/vnd.ms-excel.sheet.macroEnabled.12',
            )
            return jsonify(success=True)

        with session_scope() as session:
            projectRepository = ProjectRepository(session)
            project = projectRepository.find_by_id(project_id)
            existing_name = project.table_config_minio_id if project else None

            if existing_name:
                try:
                    self._minio.copy_to_bucket(_TC_BUCKET, _TC_BUCKET_ARCH, existing_name)
                    self._minio.delete(_TC_BUCKET, existing_name)
                except AppError:
                    pass

            self._minio.upload_bytes(
                _TC_BUCKET, object_name, content,
                content_type='application/vnd.ms-excel.sheet.macroEnabled.12',
            )

            if project:
                project.table_config_minio_id = object_name
                projectRepository.save(project)

        return jsonify(success=True)

    def download_table_config_system(self) -> Response:
        current_user = getattr(g, 'current_user', None)
        project_id = getattr(current_user, 'project_id', None) if current_user else None

        if not project_id:
            raise AppError('Проект не определён.')

        with session_scope() as session:
            project = ProjectRepository(session).find_by_id(project_id)
            if not project or not project.table_config_minio_id:
                raise AppError('Конфигурационный файл в системе отсутствует.')
            object_name = project.table_config_minio_id

        content = self._minio.download_bytes(_TC_BUCKET, object_name)
        return self._build_xlsm_response(content, f'{object_name}.xlsm')

    def validate_table_config_system(self) -> Response:
        current_user = getattr(g, 'current_user', None)
        project_id = getattr(current_user, 'project_id', None) if current_user else None

        if not project_id:
            raise AppError('Проект не определён.')

        with session_scope() as session:
            project = ProjectRepository(session).find_by_id(project_id)
            if not project or not project.table_config_minio_id:
                raise AppError('Конфигурационный файл в системе отсутствует.')
            content = self._minio.download_bytes(_TC_BUCKET, project.table_config_minio_id)

        tables = self._config_parser.parse_tables_config(content, 'config.xlsm')
        self._config_validator.validate_tables(tables)
        return jsonify(success=True)

    def validate_table_config_local(self, request: Request) -> Response:
        file = request.files.get('config_file')
        if not file or not file.filename:
            raise AppError('Файл не выбран.')
        content = file.read()
        if not content:
            raise AppError('Загруженный файл пуст.')

        tables = self._config_parser.parse_tables_config(content, file.filename)
        self._config_validator.validate_tables(tables)
        return jsonify(success=True)

    def _append_tables(
        self,
        existing_bytes: bytes,
        tables: list[tuple[str, str, list[dict]]],
        add_pk: bool,
        add_package_fields: bool,
    ) -> bytes:
        processed = [
            (name, original_name, self._apply_extra_columns(list(cols), add_pk, add_package_fields))
            for name, original_name, cols in tables
        ]

        template_bytes = _TEMPLATE_PATH.read_bytes()
        wb = load_workbook(BytesIO(existing_bytes), keep_vba=True)
        ws = wb['tables_config']
        wb.active = ws

        existing_names = {
            ws.cell(row=1 + i * _TC_BLOCK_STRIDE, column=2).value
            for i in range(_TC_MAX_BLOCKS)
            if ws.cell(row=1 + i * _TC_BLOCK_STRIDE, column=2).value is not None
        }
        duplicates = [name for name, _, _ in processed if name in existing_names]
        if duplicates:
            raise ValidationError(errors=[f'Таблица уже существует в конфигурационном файле: {", ".join(duplicates)}'])

        start_block = self._find_first_empty_block(ws)
        if start_block < 0:
            raise AppError(f'Конфигурационный файл уже заполнен (максимум {_TC_MAX_BLOCKS} таблиц).')

        available = _TC_MAX_BLOCKS - start_block
        if len(processed) > available:
            raise AppError(
                f'Недостаточно свободных блоков: доступно {available}, передано {len(processed)}.'
            )

        for i, (name, original_name, cols) in enumerate(processed):
            self._fill_tc_block(ws, 1 + (start_block + i) * _TC_BLOCK_STRIDE, name, cols, original_name)

        all_rows = _TC_BLOCK_STRIDE * _TC_MAX_BLOCKS
        label_width = max(
            (len(str(ws.cell(row=r, column=1).value or '')) for r in range(1, _TC_DATA_ROWS + 1)),
            default=0,
        )
        if label_width:
            ws.column_dimensions['A'].width = min(label_width + 4, 60)

        for col_idx in range(2, ws.max_column + 1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            col_width = max(
                (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, all_rows + 1)),
                default=0,
            )
            if col_width:
                ws.column_dimensions[letter].width = min(col_width + 4, 60)

        output = BytesIO()
        wb.save(output)
        return self._restore_x14_validations(template_bytes, output.getvalue())

    @staticmethod
    def _find_first_empty_block(ws) -> int:
        for i in range(_TC_MAX_BLOCKS):
            if ws.cell(row=1 + i * _TC_BLOCK_STRIDE, column=2).value is None:
                return i
        return -1

    def _read_and_infer(self, content: bytes, filename: str) -> tuple[str, str, list[dict]]:
        table_name, original_name, headers, rows = self._reader.read_data_file(content, filename)
        self._validator.validate_data_file(headers, rows)
        columns = self._reader.infer_columns(headers, rows)
        self._validator.validate_translated_columns(columns)
        return table_name, original_name, columns

    @staticmethod
    def _build_xlsm_response(xlsx_bytes: bytes, download_name: str) -> Response:
        ascii_name = download_name.encode('ascii', 'replace').decode('ascii')
        encoded_name = quote(download_name, encoding='utf-8')
        return Response(
            xlsx_bytes,
            mimetype='application/vnd.ms-excel.sheet.macroEnabled.12',
            headers={
                'Content-Disposition': (
                    f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
                ),
            },
        )

    def generate_excel_config(
        self,
        table_name: str,
        columns: list[dict],
        add_pk: bool = False,
        add_package_fields: bool = False,
    ) -> bytes:
        return self.generate_excel_config_multi(
            [(table_name, None, columns)], add_pk=add_pk, add_package_fields=add_package_fields
        )

    def generate_excel_config_multi(
        self,
        tables: list[tuple[str, str | None, list[dict]]],
        add_pk: bool = False,
        add_package_fields: bool = False,
    ) -> bytes:
        if not tables:
            raise AppError('Нет данных для формирования конфигурации.')
        if len(tables) > _TC_MAX_BLOCKS:
            raise AppError(
                f'Превышено максимальное количество таблиц: {len(tables)} (допустимо {_TC_MAX_BLOCKS}).'
            )

        processed: list[tuple[str, str | None, list[dict]]] = [
            (name, original_name, self._apply_extra_columns(list(cols), add_pk, add_package_fields))
            for name, original_name, cols in tables
        ]

        wb = load_workbook(_TEMPLATE_PATH, keep_vba=True)
        template_bytes = _TEMPLATE_PATH.read_bytes()
        ws = wb['tables_config']
        wb.active = ws

        clear_last_row = _TC_BLOCK_STRIDE * _TC_MAX_BLOCKS
        for row in ws.iter_rows(min_row=1, max_row=clear_last_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        max_columns = 0
        for block_idx, (name, original_name, cols) in enumerate(processed):
            start_row = 1 + block_idx * _TC_BLOCK_STRIDE
            self._fill_tc_block(ws, start_row, name, cols, original_name)
            max_columns = max(max_columns, len(cols))

        label_width = max(
            (len(str(ws.cell(row=r, column=1).value or '')) for r in range(1, _TC_DATA_ROWS + 1)),
            default=0,
        )
        if label_width:
            ws.column_dimensions['A'].width = min(label_width + 4, 60)

        used_rows = _TC_BLOCK_STRIDE * len(processed)
        for col_idx in range(2, max_columns + 2):
            letter = ws.cell(row=1, column=col_idx).column_letter
            col_width = max(
                (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, used_rows + 1)),
                default=0,
            )
            if col_width:
                ws.column_dimensions[letter].width = min(col_width + 4, 60)

        output = BytesIO()
        wb.save(output)
        return self._restore_x14_validations(template_bytes, output.getvalue())

    @staticmethod
    def _fill_tc_block(
        ws, start_row: int, table_name: str, columns: list[dict], original_name: str | None = None
    ) -> None:
        # Row offsets within block: 0=table name, 1=label, 2=code, 3=type, 4=size,
        # 5=mandatory, 6=unique, 7=pk, 8=fk, 9=default
        ws.cell(row=start_row, column=2).value = table_name
        if original_name:
            ws.cell(row=start_row, column=3).value = original_name
        for col_idx, col_info in enumerate(columns, start=2):
            ws.cell(row=start_row + 1, column=col_idx).value = col_info.get('label') or col_info['code']
            ws.cell(row=start_row + 2, column=col_idx).value = col_info['code']
            ws.cell(row=start_row + 3, column=col_idx).value = col_info['db_type']
            if col_info.get('size'):
                ws.cell(row=start_row + 4, column=col_idx).value = col_info['size']
            if col_info.get('nullable') is False:
                ws.cell(row=start_row + 5, column=col_idx).value = 'да'
            if col_info.get('unique'):
                ws.cell(row=start_row + 6, column=col_idx).value = 'да'
            if col_info.get('primary_key'):
                ws.cell(row=start_row + 7, column=col_idx).value = 'да'
            if col_info.get('foreign_key'):
                ws.cell(row=start_row + 8, column=col_idx).value = col_info['foreign_key']
            if col_info.get('default') is not None:
                ws.cell(row=start_row + 9, column=col_idx).value = col_info['default']

    def generate_excel_config_v2(
        self,
        table_name: str,
        columns: list[dict],
        add_pk: bool = False,
        add_package_fields: bool = False,
    ) -> bytes:
        columns = self._apply_extra_columns(list(columns), add_pk, add_package_fields)

        wb = load_workbook(_TEMPLATE_PATH, keep_vba=True)
        template_bytes = _TEMPLATE_PATH.read_bytes()
        ws = wb['tables_config_v2']
        wb.active = ws

        ws.cell(row=1, column=2).value = table_name
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=_V2_DATA_COLS):
            for cell in row:
                cell.value = None

        for row_idx, col_info in enumerate(columns, start=3):
            ws.cell(row=row_idx, column=1).value = col_info.get('label') or col_info['code']
            ws.cell(row=row_idx, column=2).value = col_info['code']
            ws.cell(row=row_idx, column=3).value = col_info['db_type']
            if col_info.get('size'):
                ws.cell(row=row_idx, column=4).value = col_info['size']
            if col_info.get('primary_key'):
                ws.cell(row=row_idx, column=7).value = 'да'

        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=_V2_DATA_COLS):
            max_len = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
            if max_len > 0:
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        output = BytesIO()
        wb.save(output)
        return self._restore_x14_validations(template_bytes, output.getvalue())

    @staticmethod
    def _apply_extra_columns(columns: list[dict], add_pk: bool, add_package_fields: bool) -> list[dict]:
        if add_pk and not any(c.get('primary_key') for c in columns):
            columns = [_AUTO_PK_COL] + columns

        if add_package_fields:
            existing = {c['code'].lower() for c in columns}
            need_pkg_id = 'package_id' not in existing
            need_pkg_ts = 'package_timestamp' not in existing
            if need_pkg_id or need_pkg_ts:
                ref_idx = next((i for i, c in enumerate(columns) if c['code'].lower() == 'package_id'), -1)
                if ref_idx < 0:
                    ref_idx = next((i for i, c in enumerate(columns) if c['code'].lower() == 'id'), -1)
                insert_at = ref_idx + 1 if ref_idx >= 0 else 0
                pkg_cols: list[dict] = []
                if need_pkg_id:
                    pkg_cols.append(_PACKAGE_ID_COL)
                if need_pkg_ts:
                    pkg_cols.append(_PACKAGE_TS_COL)
                columns = columns[:insert_at] + pkg_cols + columns[insert_at:]

        return columns

    @staticmethod
    def _sheet_name_to_zip_path(workbook_xml: str, rels_xml: str) -> dict[str, str]:
        name_rid: dict[str, str] = {}
        for m in re.finditer(r'<sheet\s[^>]*\bname="([^"]*)"[^>]*\br:id="([^"]*)"', workbook_xml):
            name_rid[m.group(1)] = m.group(2)
        rid_zip: dict[str, str] = {}
        for m in re.finditer(r'<Relationship\s([^>]*)>', rels_xml):
            attrs = m.group(1)
            id_m = re.search(r'\bId="([^"]*)"', attrs)
            target_m = re.search(r'\bTarget="([^"]*)"', attrs)
            if not id_m or not target_m:
                continue
            target = target_m.group(1)
            zip_path = target.lstrip('/') if target.startswith('/') else f'xl/{target}'
            rid_zip[id_m.group(1)] = zip_path
        return {name: rid_zip[rid] for name, rid in name_rid.items() if rid in rid_zip}

    def _restore_x14_validations(self, template_bytes: bytes, output_bytes: bytes) -> bytes:
        wb_rel_path = 'xl/_rels/workbook.xml.rels'
        wb_path = 'xl/workbook.xml'

        with zipfile.ZipFile(BytesIO(template_bytes)) as tz:
            tmpl_name_to_zip = self._sheet_name_to_zip_path(
                tz.read(wb_path).decode('utf-8'),
                tz.read(wb_rel_path).decode('utf-8'),
            )
            tmpl_extlst: dict[str, tuple[str, str] | None] = {}
            for name, zip_path in tmpl_name_to_zip.items():
                xml = tz.read(zip_path).decode('utf-8')
                m = re.search(r'<extLst>.*?</extLst>', xml, re.DOTALL)
                if not m:
                    tmpl_extlst[name] = None
                    continue
                root_m = re.search(r'<worksheet([^>]*)>', xml)
                tmpl_extlst[name] = (m.group(0), root_m.group(1) if root_m else '')

        out_buf = BytesIO(output_bytes)
        with zipfile.ZipFile(out_buf, 'r') as oz:
            out_name_to_zip = self._sheet_name_to_zip_path(
                oz.read(wb_path).decode('utf-8'),
                oz.read(wb_rel_path).decode('utf-8'),
            )
            out_entries = {item.filename: oz.read(item.filename) for item in oz.infolist()}

        for sheet_name, payload in tmpl_extlst.items():
            if not payload:
                continue
            extlst, tmpl_root_attrs = payload
            zip_path = out_name_to_zip.get(sheet_name)
            if not zip_path or zip_path not in out_entries:
                continue
            xml = out_entries[zip_path].decode('utf-8')
            root_tag_m = re.match(r'(<worksheet)([^>]*)(>)', xml)
            if root_tag_m:
                cur_attrs = root_tag_m.group(2)
                extra_ns = ''
                for ns_m in re.finditer(r'(xmlns:[a-zA-Z0-9_]+="[^"]*")', tmpl_root_attrs):
                    ns_decl = ns_m.group(1)
                    prefix = re.match(r'xmlns:([a-zA-Z0-9_]+)=', ns_decl).group(1)
                    if (f'{prefix}:' in extlst) and (f'xmlns:{prefix}=' not in cur_attrs):
                        extra_ns += ' ' + ns_decl
                if extra_ns:
                    xml = root_tag_m.group(1) + cur_attrs + extra_ns + root_tag_m.group(3) + xml[root_tag_m.end():]
            xml = re.sub(r'<extLst>.*?</extLst>', '', xml, flags=re.DOTALL)
            xml = xml.replace('</worksheet>', extlst + '</worksheet>', 1)
            out_entries[zip_path] = xml.encode('utf-8')

        result = BytesIO()
        with zipfile.ZipFile(result, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for filename, data in out_entries.items():
                zout.writestr(filename, data)
        return result.getvalue()
